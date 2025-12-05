"""
Generate sample Excel report for demonstration purposes.
Run this script to create the sample_report_contoso_20251204.xlsx file.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

# Create output directory
output_dir = Path(__file__).parent
output_file = output_dir / "sample_report_contoso_20251204.xlsx"

# Sample data for Executive Summary
executive_summary = pd.DataFrame({
    'Metric': [
        'Customer Name',
        'Report Period Start',
        'Report Period End',
        'Currency',
        '',
        'Total Retail (List) Cost',
        'Total Negotiated Cost',
        'Total Effective Cost',
        'Total Savings',
        'Savings Percentage',
        '',
        'Negotiated Discount Savings',
        'Reserved Instance Savings',
        'Savings Plan Savings',
        'Azure Hybrid Benefit Savings',
        'Dev/Test Pricing Savings'
    ],
    'Value': [
        'Contoso Corporation',
        '2025-09-01',
        '2025-12-01',
        'USD',
        '',
        2714850.00,
        2443365.00,
        1867616.00,
        847234.00,
        '31.2%',
        '',
        271485.00,
        312450.00,
        156780.00,
        78519.00,
        28000.00
    ]
})

# Sample data for Savings by Service
savings_by_service = pd.DataFrame({
    'ServiceCategory': [
        'Virtual Machines',
        'SQL Database',
        'Azure Kubernetes Service',
        'Storage Accounts',
        'Azure App Service',
        'Azure Cosmos DB',
        'Networking',
        'Azure Functions',
        'Azure Monitor',
        'Key Vault',
        'Azure Cache for Redis',
        'Event Hubs',
        'Service Bus',
        'API Management',
        'Other Services'
    ],
    'ListCost': [1245000, 456000, 312000, 198500, 156000, 124350, 98000, 65000, 24000, 12000, 8500, 7200, 5100, 3200, 60000],
    'BilledCost': [1120500, 410400, 280800, 178650, 140400, 111915, 88200, 58500, 21600, 10800, 7650, 6480, 4590, 2880, 54000],
    'EffectiveCost': [785000, 298000, 234000, 156200, 112000, 98416, 78000, 52000, 18000, 9000, 6800, 5600, 4000, 2600, 54000],
    'NegotiatedSavings': [124500, 45600, 31200, 19850, 15600, 12435, 9800, 6500, 2400, 1200, 850, 720, 510, 320, 6000],
    'CommitmentSavings': [335500, 112400, 46800, 22450, 28400, 13499, 10200, 6500, 3600, 1800, 850, 880, 590, 280, 0],
    'TotalSavings': [460000, 158000, 78000, 42300, 44000, 25934, 20000, 13000, 6000, 3000, 1700, 1600, 1100, 600, 6000],
    'SavingsPercentage': [36.9, 34.6, 25.0, 21.3, 28.2, 20.9, 20.4, 20.0, 25.0, 25.0, 20.0, 22.2, 21.6, 18.8, 10.0]
})

# Sample data for Savings by Subscription
savings_by_subscription = pd.DataFrame({
    'SubAccountId': [
        'sub-001-prod',
        'sub-002-staging',
        'sub-003-dev',
        'sub-004-shared'
    ],
    'SubAccountName': [
        'Contoso-Production',
        'Contoso-Staging',
        'Contoso-Development',
        'Contoso-Shared-Services'
    ],
    'ListCost': [1856000, 412000, 246850, 200000],
    'BilledCost': [1670400, 370800, 222165, 180000],
    'EffectiveCost': [1245000, 298000, 174616, 150000],
    'NegotiatedSavings': [185600, 41200, 24685, 20000],
    'CommitmentSavings': [425400, 72800, 47549, 30000],
    'TotalSavings': [611000, 114000, 72234, 50000],
    'SavingsPercentage': [32.9, 27.7, 29.3, 25.0]
})

# Sample data for Monthly Trend
monthly_trend = pd.DataFrame({
    'Month': ['2025-09', '2025-10', '2025-11'],
    'ListCost': [892450, 915200, 907200],
    'BilledCost': [803205, 823680, 816480],
    'EffectiveCost': [612000, 635616, 620000],
    'NegotiatedSavings': [89245, 91520, 90720],
    'CommitmentSavings': [191205, 188064, 196480],
    'TotalSavings': [280450, 279584, 287200],
    'SavingsPercentage': [31.4, 30.5, 31.7]
})

# Sample data for Top Resources
top_resources = pd.DataFrame({
    'ResourceId': [
        '/subscriptions/sub-001/resourceGroups/rg-prod/providers/Microsoft.Sql/servers/prod-sql-primary',
        '/subscriptions/sub-001/resourceGroups/rg-prod/providers/Microsoft.ContainerService/managedClusters/prod-aks-cluster-01',
        '/subscriptions/sub-001/resourceGroups/rg-prod/providers/Microsoft.Compute/virtualMachines/prod-vm-app-server-01',
        '/subscriptions/sub-001/resourceGroups/rg-prod/providers/Microsoft.Compute/virtualMachineScaleSets/prod-vm-web-cluster',
        '/subscriptions/sub-001/resourceGroups/rg-prod/providers/Microsoft.DocumentDB/databaseAccounts/prod-cosmos-orders',
        '/subscriptions/sub-003/resourceGroups/rg-dev/providers/Microsoft.Compute/virtualMachineScaleSets/dev-vm-pool',
        '/subscriptions/sub-001/resourceGroups/rg-prod/providers/Microsoft.Sql/servers/prod-sql-analytics',
        '/subscriptions/sub-001/resourceGroups/rg-prod/providers/Microsoft.Web/serverFarms/prod-app-service-plan',
        '/subscriptions/sub-001/resourceGroups/rg-prod/providers/Microsoft.Cache/Redis/prod-redis-cache',
        '/subscriptions/sub-002/resourceGroups/rg-staging/providers/Microsoft.Compute/virtualMachines/staging-vm-01'
    ],
    'ResourceName': [
        'prod-sql-primary',
        'prod-aks-cluster-01',
        'prod-vm-app-server-01',
        'prod-vm-web-cluster',
        'prod-cosmos-orders',
        'dev-vm-pool',
        'prod-sql-analytics',
        'prod-app-service-plan',
        'prod-redis-cache',
        'staging-vm-01'
    ],
    'ServiceCategory': [
        'SQL Database',
        'Azure Kubernetes Service',
        'Virtual Machines',
        'Virtual Machines',
        'Azure Cosmos DB',
        'Virtual Machines',
        'SQL Database',
        'Azure App Service',
        'Azure Cache for Redis',
        'Virtual Machines'
    ],
    'SavingsCategory': [
        'Reserved Instance',
        'Savings Plan',
        'RI + AHB',
        'Reserved Instance',
        'Negotiated Discount',
        'Dev/Test Pricing',
        'RI + AHB',
        'Savings Plan',
        'Reserved Instance',
        'Savings Plan'
    ],
    'ListCost': [98000, 78500, 65400, 58900, 45600, 38000, 35200, 28600, 18500, 22400],
    'EffectiveCost': [52400, 40300, 33300, 30000, 27150, 22800, 20400, 16000, 12200, 14800],
    'TotalSavings': [45600, 38200, 32100, 28900, 18450, 15200, 14800, 12600, 6300, 7600],
    'SavingsPercentage': [46.5, 48.7, 49.1, 49.1, 40.5, 40.0, 42.0, 44.1, 34.1, 33.9]
})

# Sample data for Reserved Instances
reserved_instances = pd.DataFrame({
    'ReservationId': ['ri-001', 'ri-002', 'ri-003', 'ri-004', 'ri-005'],
    'ReservationName': [
        'RI-Prod-VMs-EastUS',
        'RI-SQL-Premium',
        'RI-Prod-VMs-WestUS',
        'RI-CosmosDB-Reserved',
        'RI-Redis-Cache'
    ],
    'ServiceCategory': [
        'Virtual Machines',
        'SQL Database',
        'Virtual Machines',
        'Azure Cosmos DB',
        'Azure Cache for Redis'
    ],
    'Term': ['3 Year', '3 Year', '1 Year', '1 Year', '1 Year'],
    'QuantityPurchased': [50, 20, 25, 10, 5],
    'QuantityUsed': [47.1, 17.7, 22.8, 8.5, 4.2],
    'Utilization': ['94.2%', '88.5%', '91.3%', '85.0%', '84.0%'],
    'ListCost': [312000, 178900, 134000, 45000, 18500],
    'EffectiveCost': [156000, 89450, 67000, 24500, 12200],
    'Savings': [156000, 89450, 67000, 20500, 6300]
})

# Sample data for Savings Plans
savings_plans = pd.DataFrame({
    'SavingsPlanId': ['sp-001', 'sp-002'],
    'SavingsPlanName': ['SP-Compute-3Year', 'SP-Compute-1Year'],
    'Term': ['3 Year', '1 Year'],
    'HourlyCommitment': [68.49, 34.25],
    'MonthlyCommitment': [50000, 25000],
    'Utilization': ['96.8%', '92.1%'],
    'CoveredServices': ['VMs, AKS, App Service, Functions', 'VMs, AKS'],
    'ListCostCovered': [245000, 142000],
    'EffectiveCost': [146220, 84000],
    'Savings': [98780, 58000]
})

# Sample data for Azure Hybrid Benefit
ahb_savings = pd.DataFrame({
    'LicenseType': ['Windows Server', 'SQL Server Standard', 'SQL Server Enterprise'],
    'ResourceCount': [45, 12, 5],
    'ListCost': [89500, 56000, 42000],
    'EffectiveCost': [45000, 35981, 28000],
    'Savings': [44500, 20019, 14000],
    'SavingsPercentage': ['49.7%', '35.7%', '33.3%']
})

# Write to Excel with multiple sheets
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    executive_summary.to_excel(writer, sheet_name='Executive Summary', index=False)
    savings_by_service.to_excel(writer, sheet_name='By Service', index=False)
    savings_by_subscription.to_excel(writer, sheet_name='By Subscription', index=False)
    monthly_trend.to_excel(writer, sheet_name='Monthly Trend', index=False)
    top_resources.to_excel(writer, sheet_name='Top Resources', index=False)
    reserved_instances.to_excel(writer, sheet_name='Reserved Instances', index=False)
    savings_plans.to_excel(writer, sheet_name='Savings Plans', index=False)
    ahb_savings.to_excel(writer, sheet_name='Azure Hybrid Benefit', index=False)

# Now add charts to the workbook
wb = load_workbook(output_file)

# Create Charts sheet
charts_ws = wb.create_sheet('Charts', 0)  # Insert at the beginning

# Add chart data for savings distribution pie chart
chart_data_start_row = 20  # Place data below the charts
charts_ws.cell(row=chart_data_start_row, column=1, value='Savings Category')
charts_ws.cell(row=chart_data_start_row, column=2, value='Amount')

savings_categories = [
    ('Negotiated Discount', 271485),
    ('Reserved Instances', 312450),
    ('Savings Plans', 156780),
    ('Azure Hybrid Benefit', 78519),
    ('Dev/Test Pricing', 28000)
]

for i, (category, amount) in enumerate(savings_categories, start=1):
    charts_ws.cell(row=chart_data_start_row + i, column=1, value=category)
    charts_ws.cell(row=chart_data_start_row + i, column=2, value=amount)

# Create Pie Chart for Savings Distribution
pie = PieChart()
pie.title = "Savings by Category"
labels = Reference(charts_ws, min_col=1, min_row=chart_data_start_row + 1, max_row=chart_data_start_row + 5)
data = Reference(charts_ws, min_col=2, min_row=chart_data_start_row, max_row=chart_data_start_row + 5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.width = 15
pie.height = 10

# Add data labels to pie chart
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
pie.dataLabels.showVal = False

charts_ws.add_chart(pie, "A1")

# Add monthly trend data for bar chart
trend_data_start_row = 30
charts_ws.cell(row=trend_data_start_row, column=1, value='Month')
charts_ws.cell(row=trend_data_start_row, column=2, value='Savings ($)')

monthly_savings = [
    ('Sep 2025', 267500),
    ('Oct 2025', 285000),
    ('Nov 2025', 294734)
]

for i, (month, savings) in enumerate(monthly_savings, start=1):
    charts_ws.cell(row=trend_data_start_row + i, column=1, value=month)
    charts_ws.cell(row=trend_data_start_row + i, column=2, value=savings)

# Create Bar Chart for Monthly Trend
bar = BarChart()
bar.title = "Monthly Savings Trend"
bar.type = "col"
bar.style = 10
bar.y_axis.title = "Savings ($)"
bar.x_axis.title = "Month"

bar_data = Reference(charts_ws, min_col=2, min_row=trend_data_start_row, max_row=trend_data_start_row + 3)
bar_cats = Reference(charts_ws, min_col=1, min_row=trend_data_start_row + 1, max_row=trend_data_start_row + 3)
bar.add_data(bar_data, titles_from_data=True)
bar.set_categories(bar_cats)
bar.width = 15
bar.height = 10
bar.shape = 4

charts_ws.add_chart(bar, "J1")

# Add summary text
charts_ws.cell(row=1, column=20, value="Total Savings: $847,234")
charts_ws.cell(row=2, column=20, value="Savings Rate: 31.2%")
charts_ws.cell(row=3, column=20, value="Report Period: Sep-Nov 2025")

# Save the workbook with charts
wb.save(output_file)

print(f"Sample Excel report generated: {output_file}")
print("Charts sheet added with Pie Chart (Savings by Category) and Bar Chart (Monthly Trend)")
