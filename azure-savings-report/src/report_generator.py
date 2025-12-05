"""
Azure Savings Report - Report Generator
Generates HTML and Excel reports from savings data
"""

import os
from datetime import datetime
from pathlib import Path
from typing import Optional
import pandas as pd

from .savings_calculator import SavingsReport


class ReportGenerator:
    """Generates formatted reports from savings data"""
    
    def __init__(self, output_dir: str = "./reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def _format_currency(self, value: float, currency: str = "USD") -> str:
        """Format a value as currency"""
        symbols = {"USD": "$", "EUR": "€", "GBP": "£"}
        symbol = symbols.get(currency, currency + " ")
        return f"{symbol}{value:,.2f}"
    
    def _format_percentage(self, value: float) -> str:
        """Format a value as percentage"""
        return f"{value:.1f}%"
    
    def generate_html_report(self, report: SavingsReport) -> str:
        """Generate an HTML report"""
        
        html_template = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Azure Savings Report - {report.customer_name}</title>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        :root {{
            --primary-color: #0078d4;
            --secondary-color: #50e6ff;
            --success-color: #107c10;
            --background-color: #f3f2f1;
            --card-background: #ffffff;
            --text-color: #323130;
            --border-color: #edebe9;
        }}
        
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: var(--background-color);
            color: var(--text-color);
            line-height: 1.6;
        }}
        
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
        }}
        
        header {{
            background: linear-gradient(135deg, var(--primary-color), #005a9e);
            color: white;
            padding: 40px;
            margin-bottom: 30px;
            border-radius: 8px;
        }}
        
        header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
        }}
        
        header .subtitle {{
            font-size: 1.2em;
            opacity: 0.9;
        }}
        
        .summary-cards {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        
        .card {{
            background: var(--card-background);
            border-radius: 8px;
            padding: 25px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        
        .card h3 {{
            color: var(--primary-color);
            margin-bottom: 15px;
            font-size: 1.1em;
        }}
        
        .card .value {{
            font-size: 2em;
            font-weight: 600;
            color: var(--text-color);
        }}
        
        .card .value.savings {{
            color: var(--success-color);
        }}
        
        .card .label {{
            color: #605e5c;
            font-size: 0.9em;
            margin-top: 5px;
        }}
        
        .section {{
            background: var(--card-background);
            border-radius: 8px;
            padding: 30px;
            margin-bottom: 30px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        
        .section h2 {{
            color: var(--primary-color);
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid var(--border-color);
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }}
        
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid var(--border-color);
        }}
        
        th {{
            background-color: #f3f2f1;
            font-weight: 600;
        }}
        
        tr:hover {{
            background-color: #f9f9f9;
        }}
        
        .savings-breakdown {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-top: 20px;
        }}
        
        .breakdown-item {{
            background: #f9f9f9;
            padding: 15px;
            border-radius: 6px;
            border-left: 4px solid var(--primary-color);
        }}
        
        .breakdown-item.ri {{ border-left-color: #0078d4; }}
        .breakdown-item.sp {{ border-left-color: #00b294; }}
        .breakdown-item.ahb {{ border-left-color: #8764b8; }}
        .breakdown-item.devtest {{ border-left-color: #ff8c00; }}
        .breakdown-item.negotiated {{ border-left-color: #107c10; }}
        
        .chart-container {{
            width: 100%;
            height: 400px;
            margin: 20px 0;
        }}
        
        footer {{
            text-align: center;
            padding: 20px;
            color: #605e5c;
            font-size: 0.9em;
        }}
        
        @media print {{
            body {{ background: white; }}
            .card, .section {{ box-shadow: none; border: 1px solid #ddd; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>Azure Savings Realization Report</h1>
            <div class="subtitle">{report.customer_name}</div>
            <div class="subtitle">Period: {report.report_period_start.strftime('%B %d, %Y')} - {report.report_period_end.strftime('%B %d, %Y')}</div>
        </header>
        
        <!-- Executive Summary -->
        <div class="summary-cards">
            <div class="card">
                <h3>Retail (List) Cost</h3>
                <div class="value">{self._format_currency(report.total_retail_cost, report.currency)}</div>
                <div class="label">What you would pay at public pricing</div>
            </div>
            <div class="card">
                <h3>Your Effective Cost</h3>
                <div class="value">{self._format_currency(report.total_effective_cost, report.currency)}</div>
                <div class="label">What you actually paid</div>
            </div>
            <div class="card">
                <h3>Total Savings</h3>
                <div class="value savings">{self._format_currency(report.total_savings, report.currency)}</div>
                <div class="label">{self._format_percentage(report.total_savings_percentage)} savings realized</div>
            </div>
        </div>
        
        <!-- Savings Breakdown -->
        <div class="section">
            <h2>Savings Breakdown by Category</h2>
            <div class="savings-breakdown">
                <div class="breakdown-item negotiated">
                    <h4>Negotiated Discount</h4>
                    <div class="value">{self._format_currency(report.negotiated_discount_savings.total_savings, report.currency)}</div>
                    <div class="label">EA/MCA contract pricing vs retail</div>
                </div>
                <div class="breakdown-item ri">
                    <h4>Reserved Instances</h4>
                    <div class="value">{self._format_currency(report.reserved_instance_savings.total_savings, report.currency)}</div>
                    <div class="label">{self._format_percentage(report.reserved_instance_savings.savings_percentage)} discount</div>
                </div>
                <div class="breakdown-item sp">
                    <h4>Savings Plans</h4>
                    <div class="value">{self._format_currency(report.savings_plan_savings.total_savings, report.currency)}</div>
                    <div class="label">{self._format_percentage(report.savings_plan_savings.savings_percentage)} discount</div>
                </div>
                <div class="breakdown-item ahb">
                    <h4>Azure Hybrid Benefit</h4>
                    <div class="value">{self._format_currency(report.ahb_savings.total_savings, report.currency)}</div>
                    <div class="label">Windows/SQL Server license savings</div>
                </div>
                <div class="breakdown-item devtest">
                    <h4>Dev/Test Pricing</h4>
                    <div class="value">{self._format_currency(report.devtest_savings.total_savings, report.currency)}</div>
                    <div class="label">Dev/Test subscription discounts</div>
                </div>
            </div>
            
            <div id="savingsPieChart" class="chart-container"></div>
        </div>
        
        <!-- Savings by Service -->
        <div class="section">
            <h2>Savings by Service Category</h2>
            <table>
                <thead>
                    <tr>
                        <th>Service</th>
                        <th>Retail Cost</th>
                        <th>Effective Cost</th>
                        <th>Savings</th>
                        <th>Savings %</th>
                    </tr>
                </thead>
                <tbody>
                    {self._generate_service_table_rows(report)}
                </tbody>
            </table>
        </div>
        
        <!-- Monthly Trend -->
        <div class="section">
            <h2>Monthly Savings Trend</h2>
            <div id="monthlyTrendChart" class="chart-container"></div>
        </div>
        
        <!-- Top Savings Resources -->
        <div class="section">
            <h2>Top Resources by Savings</h2>
            <table>
                <thead>
                    <tr>
                        <th>Resource</th>
                        <th>Service</th>
                        <th>Savings Category</th>
                        <th>Savings</th>
                    </tr>
                </thead>
                <tbody>
                    {self._generate_resources_table_rows(report)}
                </tbody>
            </table>
        </div>
        
        <footer>
            <p>Generated on {datetime.now().strftime('%B %d, %Y at %H:%M')} | Azure Savings Report Tool</p>
        </footer>
    </div>
    
    <script>
        // Savings Pie Chart
        var pieData = [{{
            values: [
                {report.negotiated_discount_savings.total_savings},
                {report.reserved_instance_savings.total_savings},
                {report.savings_plan_savings.total_savings},
                {report.ahb_savings.total_savings},
                {report.devtest_savings.total_savings}
            ],
            labels: ['Negotiated Discount', 'Reserved Instances', 'Savings Plans', 'Azure Hybrid Benefit', 'Dev/Test Pricing'],
            type: 'pie',
            hole: 0.4,
            marker: {{
                colors: ['#107c10', '#0078d4', '#00b294', '#8764b8', '#ff8c00']
            }}
        }}];
        
        var pieLayout = {{
            title: 'Savings Distribution by Category',
            showlegend: true,
            legend: {{ orientation: 'h', y: -0.1 }}
        }};
        
        Plotly.newPlot('savingsPieChart', pieData, pieLayout, {{responsive: true}});
        
        // Monthly Trend Chart
        {self._generate_monthly_chart_data(report)}
    </script>
</body>
</html>
"""
        
        # Save the report
        filename = f"savings_report_{report.customer_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.html"
        filepath = self.output_dir / filename
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_template)
        
        return str(filepath)
    
    def _generate_service_table_rows(self, report: SavingsReport) -> str:
        """Generate table rows for services"""
        rows = []
        for _, row in report.savings_by_service.head(15).iterrows():
            rows.append(f"""
                <tr>
                    <td>{row['ServiceCategory']}</td>
                    <td>{self._format_currency(row['ListCost'], report.currency)}</td>
                    <td>{self._format_currency(row['EffectiveCost'], report.currency)}</td>
                    <td>{self._format_currency(row['TotalSavings'], report.currency)}</td>
                    <td>{self._format_percentage(row['SavingsPercentage'])}</td>
                </tr>
            """)
        return '\n'.join(rows)
    
    def _generate_resources_table_rows(self, report: SavingsReport) -> str:
        """Generate table rows for top resources"""
        rows = []
        for _, row in report.top_savings_resources.head(10).iterrows():
            resource_name = row['ResourceName'] if len(str(row['ResourceName'])) < 50 else str(row['ResourceName'])[:47] + '...'
            rows.append(f"""
                <tr>
                    <td>{resource_name}</td>
                    <td>{row['ServiceCategory']}</td>
                    <td>{row['SavingsCategory']}</td>
                    <td>{self._format_currency(row['TotalSavings'], report.currency)}</td>
                </tr>
            """)
        return '\n'.join(rows)
    
    def _generate_monthly_chart_data(self, report: SavingsReport) -> str:
        """Generate JavaScript for monthly trend chart"""
        if report.monthly_trend.empty:
            return "// No monthly data available"
        
        months = report.monthly_trend['Month'].tolist()
        retail = report.monthly_trend['ListCost'].tolist()
        effective = report.monthly_trend['EffectiveCost'].tolist()
        savings = report.monthly_trend['TotalSavings'].tolist()
        
        return f"""
        var monthlyData = [
            {{
                x: {months},
                y: {retail},
                name: 'Retail Cost',
                type: 'bar',
                marker: {{ color: '#d13438' }}
            }},
            {{
                x: {months},
                y: {effective},
                name: 'Effective Cost',
                type: 'bar',
                marker: {{ color: '#0078d4' }}
            }},
            {{
                x: {months},
                y: {savings},
                name: 'Savings',
                type: 'scatter',
                mode: 'lines+markers',
                yaxis: 'y2',
                marker: {{ color: '#107c10' }}
            }}
        ];
        
        var monthlyLayout = {{
            title: 'Monthly Cost and Savings Trend',
            barmode: 'group',
            yaxis: {{ title: 'Cost ({report.currency})' }},
            yaxis2: {{
                title: 'Savings ({report.currency})',
                overlaying: 'y',
                side: 'right'
            }},
            legend: {{ orientation: 'h', y: -0.2 }}
        }};
        
        Plotly.newPlot('monthlyTrendChart', monthlyData, monthlyLayout, {{responsive: true}});
        """
    
    def generate_excel_report(self, report: SavingsReport) -> str:
        """Generate an Excel report with charts"""
        from openpyxl.chart import PieChart, BarChart, LineChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.series import DataPoint
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
        
        filename = f"savings_report_{report.customer_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Executive Summary
            summary_data = {
                'Metric': [
                    'Customer Name',
                    'Report Period Start',
                    'Report Period End',
                    'Currency',
                    '',
                    'Total Retail (List) Cost',
                    'Total Negotiated Cost',
                    'Total Effective Cost',
                    'Total Savings',
                    'Savings Percentage',
                    '',
                    'Negotiated Discount Savings',
                    'Reserved Instance Savings',
                    'Savings Plan Savings',
                    'Azure Hybrid Benefit Savings',
                    'Dev/Test Pricing Savings'
                ],
                'Value': [
                    report.customer_name,
                    report.report_period_start.strftime('%Y-%m-%d'),
                    report.report_period_end.strftime('%Y-%m-%d'),
                    report.currency,
                    '',
                    report.total_retail_cost,
                    report.total_negotiated_cost,
                    report.total_effective_cost,
                    report.total_savings,
                    f"{report.total_savings_percentage:.1f}%",
                    '',
                    report.negotiated_discount_savings.total_savings,
                    report.reserved_instance_savings.total_savings,
                    report.savings_plan_savings.total_savings,
                    report.ahb_savings.total_savings,
                    report.devtest_savings.total_savings
                ]
            }
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Executive Summary', index=False)
            
            # Savings by Service
            report.savings_by_service.to_excel(writer, sheet_name='By Service', index=False)
            
            # Savings by Subscription
            report.savings_by_subscription.to_excel(writer, sheet_name='By Subscription', index=False)
            
            # Monthly Trend
            if not report.monthly_trend.empty:
                report.monthly_trend.to_excel(writer, sheet_name='Monthly Trend', index=False)
            
            # Top Resources
            report.top_savings_resources.to_excel(writer, sheet_name='Top Resources', index=False)
            
            # Create Charts sheet with data and visualizations
            self._create_charts_sheet(writer, report)
        
        return str(filepath)
    
    def _create_charts_sheet(self, writer, report: SavingsReport):
        """Create a Charts sheet with visualizations similar to HTML report"""
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.series import DataPoint
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        workbook = writer.book
        ws = workbook.create_sheet('Charts')
        
        # Define styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
        title_font = Font(bold=True, size=16, color="0078D4")
        money_font = Font(bold=True, size=12)
        savings_font = Font(bold=True, size=14, color="107C10")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ============ EXECUTIVE SUMMARY SECTION ============
        ws['A1'] = 'AZURE SAVINGS REALIZATION REPORT'
        ws['A1'].font = Font(bold=True, size=20, color="0078D4")
        ws.merge_cells('A1:F1')
        
        ws['A2'] = report.customer_name
        ws['A2'].font = Font(size=14)
        
        ws['A3'] = f"Period: {report.report_period_start.strftime('%B %d, %Y')} - {report.report_period_end.strftime('%B %d, %Y')}"
        ws['A3'].font = Font(size=12, italic=True)
        
        # Summary Cards Row
        row = 5
        cards = [
            ('Retail (List) Cost', report.total_retail_cost, 'D13438'),
            ('Negotiated Cost', report.total_negotiated_cost, '0078D4'),
            ('Effective Cost', report.total_effective_cost, '0078D4'),
            ('Total Savings', report.total_savings, '107C10'),
        ]
        
        for col_idx, (label, value, color) in enumerate(cards, start=1):
            col = get_column_letter(col_idx * 2 - 1)
            ws[f'{col}{row}'] = label
            ws[f'{col}{row}'].font = Font(bold=True, size=10)
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
            
            ws[f'{col}{row+1}'] = f"${value:,.0f}"
            ws[f'{col}{row+1}'].font = Font(bold=True, size=14, color=color)
            ws[f'{col}{row+1}'].alignment = Alignment(horizontal='center')
        
        # Savings percentage
        ws['I5'] = 'Savings %'
        ws['I5'].font = Font(bold=True, size=10)
        ws['I5'].alignment = Alignment(horizontal='center')
        ws['I6'] = f"{report.total_savings_percentage:.1f}%"
        ws['I6'].font = Font(bold=True, size=14, color="107C10")
        ws['I6'].alignment = Alignment(horizontal='center')
        
        # ============ SAVINGS BY CATEGORY (PIE CHART DATA) ============
        row = 9
        ws[f'A{row}'] = 'SAVINGS BY CATEGORY'
        ws[f'A{row}'].font = title_font
        
        # Category data for pie chart
        categories = [
            ('Negotiated Discount', report.negotiated_discount_savings.total_savings, '107C10'),
            ('Reserved Instances', report.reserved_instance_savings.total_savings, '0078D4'),
            ('Savings Plans', report.savings_plan_savings.total_savings, '00B294'),
            ('Azure Hybrid Benefit', report.ahb_savings.total_savings, '8764B8'),
            ('Dev/Test Pricing', report.devtest_savings.total_savings, 'FF8C00'),
        ]
        
        # Write category data
        row = 11
        ws[f'A{row}'] = 'Category'
        ws[f'B{row}'] = 'Savings'
        ws[f'C{row}'] = 'Percentage'
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        ws[f'C{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'C{row}'].fill = header_fill
        
        total_cat_savings = sum(c[1] for c in categories)
        for i, (name, value, color) in enumerate(categories, start=1):
            r = row + i
            ws[f'A{r}'] = name
            ws[f'B{r}'] = value
            pct = (value / total_cat_savings * 100) if total_cat_savings > 0 else 0
            ws[f'C{r}'] = f"{pct:.1f}%"
            ws[f'A{r}'].border = thin_border
            ws[f'B{r}'].border = thin_border
            ws[f'C{r}'].border = thin_border
            ws[f'B{r}'].number_format = '$#,##0'
        
        # Create Pie Chart
        pie = PieChart()
        pie.title = "Savings Distribution by Category"
        labels = Reference(ws, min_col=1, min_row=row+1, max_row=row+5)
        data = Reference(ws, min_col=2, min_row=row, max_row=row+5)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 15
        pie.height = 10
        
        # Set custom colors for pie slices
        colors = ['107C10', '0078D4', '00B294', '8764B8', 'FF8C00']
        for i, color in enumerate(colors):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            pie.series[0].data_points.append(pt)
        
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        pie.dataLabels.showVal = False
        
        ws.add_chart(pie, "E10")
        
        # ============ SAVINGS BY SERVICE (BAR CHART) ============
        row = 23
        ws[f'A{row}'] = 'TOP SERVICES BY SAVINGS'
        ws[f'A{row}'].font = title_font
        
        # Write service data (top 10)
        row = 25
        ws[f'A{row}'] = 'Service'
        ws[f'B{row}'] = 'Retail Cost'
        ws[f'C{row}'] = 'Effective Cost'
        ws[f'D{row}'] = 'Savings'
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
        
        top_services = report.savings_by_service.head(10)
        for i, (_, svc_row) in enumerate(top_services.iterrows(), start=1):
            r = row + i
            ws[f'A{r}'] = svc_row.get('ServiceCategory', 'Unknown')
            ws[f'B{r}'] = svc_row.get('ListCost', 0)
            ws[f'C{r}'] = svc_row.get('EffectiveCost', 0)
            ws[f'D{r}'] = svc_row.get('TotalSavings', 0)
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{r}'].border = thin_border
            ws[f'B{r}'].number_format = '$#,##0'
            ws[f'C{r}'].number_format = '$#,##0'
            ws[f'D{r}'].number_format = '$#,##0'
        
        # Create Bar Chart for services
        bar = BarChart()
        bar.type = "col"
        bar.title = "Savings by Service"
        bar.y_axis.title = "Amount ($)"
        bar.x_axis.title = "Service"
        bar.width = 18
        bar.height = 10
        
        data_end_row = row + min(len(top_services), 10)
        data = Reference(ws, min_col=4, min_row=row, max_row=data_end_row)
        cats = Reference(ws, min_col=1, min_row=row+1, max_row=data_end_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.shape = 4
        bar.series[0].graphicalProperties.solidFill = "107C10"
        
        ws.add_chart(bar, "F24")
        
        # ============ MONTHLY TREND (IF DATA EXISTS) ============
        if not report.monthly_trend.empty:
            row = 40
            ws[f'A{row}'] = 'MONTHLY TREND'
            ws[f'A{row}'].font = title_font
            
            # Write monthly data
            row = 42
            ws[f'A{row}'] = 'Month'
            ws[f'B{row}'] = 'Retail Cost'
            ws[f'C{row}'] = 'Effective Cost'
            ws[f'D{row}'] = 'Savings'
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].fill = header_fill
            
            for i, (_, m_row) in enumerate(report.monthly_trend.iterrows(), start=1):
                r = row + i
                ws[f'A{r}'] = m_row.get('Month', '')
                ws[f'B{r}'] = m_row.get('ListCost', 0)
                ws[f'C{r}'] = m_row.get('EffectiveCost', 0)
                ws[f'D{r}'] = m_row.get('TotalSavings', 0)
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{r}'].border = thin_border
                ws[f'B{r}'].number_format = '$#,##0'
                ws[f'C{r}'].number_format = '$#,##0'
                ws[f'D{r}'].number_format = '$#,##0'
            
            # Create combined bar + line chart for monthly trend
            monthly_bar = BarChart()
            monthly_bar.type = "col"
            monthly_bar.title = "Monthly Cost and Savings Trend"
            monthly_bar.y_axis.title = "Cost ($)"
            monthly_bar.width = 18
            monthly_bar.height = 10
            
            data_end_row = row + len(report.monthly_trend)
            
            # Retail and Effective cost bars
            retail_data = Reference(ws, min_col=2, min_row=row, max_row=data_end_row)
            effective_data = Reference(ws, min_col=3, min_row=row, max_row=data_end_row)
            cats = Reference(ws, min_col=1, min_row=row+1, max_row=data_end_row)
            
            monthly_bar.add_data(retail_data, titles_from_data=True)
            monthly_bar.add_data(effective_data, titles_from_data=True)
            monthly_bar.set_categories(cats)
            
            monthly_bar.series[0].graphicalProperties.solidFill = "D13438"
            monthly_bar.series[1].graphicalProperties.solidFill = "0078D4"
            
            ws.add_chart(monthly_bar, "F41")
        
        # ============ ADJUST COLUMN WIDTHS ============
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
